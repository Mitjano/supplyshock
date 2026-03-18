"""World Bank Pink Sheet commodity price ingestion.

Downloads the World Bank Commodity Markets Outlook (CMO) Pink Sheet Excel file
and extracts monthly fertilizer prices.
Called by Celery beat task `ingest_world_bank_prices` weekly.
"""

import io
import logging
from datetime import datetime, timezone

import psycopg2
import requests
from openpyxl import load_workbook

from config import settings

logger = logging.getLogger("world_bank_prices")

# World Bank CMO Pink Sheet Excel URL
PINK_SHEET_URL = (
    "https://thedocs.worldbank.org/en/doc/"
    "5d903e848db1d1b83e0ec8f744e55570-0350012021/related/"
    "CMO-Historical-Data-Monthly.xlsx"
)

# Column header → commodity name for fertilizers
FERTILIZER_COLUMNS: dict[str, str] = {
    "Urea": "urea",
    "DAP": "dap",
    "Potassium chloride": "potash",
    "Phosphate rock": "phosphate_rock",
    # Alternative header spellings
    "TSP": "tsp",
    "Potash": "potash",
}


def ingest_world_bank_prices() -> dict:
    """Download World Bank Pink Sheet and extract fertilizer prices."""
    logger.info("Downloading World Bank Pink Sheet")

    try:
        resp = requests.get(PINK_SHEET_URL, timeout=60)
        if resp.status_code != 200:
            logger.error("World Bank download returned %d", resp.status_code)
            return {"error": f"HTTP {resp.status_code}"}
    except requests.RequestException as e:
        logger.error("World Bank download failed: %s", e)
        return {"error": str(e)}

    try:
        wb = load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)
    except Exception as e:
        logger.error("Failed to parse Excel file: %s", e)
        return {"error": str(e)}

    # Find the "Monthly Prices" sheet (or similar name)
    sheet = None
    for name in wb.sheetnames:
        if "monthly" in name.lower() and "price" in name.lower():
            sheet = wb[name]
            break
    if sheet is None:
        # Fallback: try first sheet
        sheet = wb.active
        logger.warning("Could not find 'Monthly Prices' sheet, using active sheet: %s", sheet.title)

    # Parse header row to find fertilizer columns
    header_row = None
    col_map: dict[int, str] = {}  # column index → commodity name
    date_col: int | None = None

    for row_idx, row in enumerate(sheet.iter_rows(max_row=10, values_only=False), start=1):
        for cell in row:
            val = str(cell.value).strip() if cell.value else ""
            if val.lower() in ("month", "year", "date", "period"):
                date_col = cell.column - 1  # 0-based
                header_row = row_idx
            for header, commodity in FERTILIZER_COLUMNS.items():
                if header.lower() in val.lower():
                    col_map[cell.column - 1] = commodity
        if header_row and col_map:
            break

    if not col_map:
        logger.error("Could not locate fertilizer columns in Pink Sheet")
        return {"error": "no fertilizer columns found"}

    if date_col is None:
        date_col = 0  # assume first column

    logger.info("Found %d fertilizer columns, date_col=%d, header_row=%d", len(col_map), date_col, header_row or 0)

    # Extract data rows — only take last 6 months to avoid massive inserts
    conn = psycopg2.connect(settings.DATABASE_URL_SYNC)
    try:
        inserted = 0
        skipped = 0
        rows_processed = 0

        data_rows = list(sheet.iter_rows(min_row=(header_row or 1) + 1, values_only=True))
        # Take only the last 6 rows (months)
        recent_rows = data_rows[-6:] if len(data_rows) > 6 else data_rows

        with conn.cursor() as cur:
            for row in recent_rows:
                date_val = row[date_col] if date_col < len(row) else None
                if date_val is None:
                    continue

                # Parse the date — World Bank uses various formats
                ts = _parse_wb_date(date_val)
                if ts is None:
                    skipped += 1
                    continue

                rows_processed += 1

                for col_idx, commodity in col_map.items():
                    if col_idx >= len(row):
                        continue
                    price_val = row[col_idx]
                    if price_val is None or price_val == "" or price_val == "..":
                        skipped += 1
                        continue

                    try:
                        price = float(price_val)
                    except (ValueError, TypeError):
                        skipped += 1
                        continue

                    if price <= 0:
                        skipped += 1
                        continue

                    cur.execute(
                        """
                        INSERT INTO commodity_prices (time, commodity, benchmark, price, currency, source)
                        VALUES (%s, %s, %s, %s, 'USD', 'world_bank')
                        ON CONFLICT (time, commodity, benchmark) DO NOTHING
                        """,
                        (ts, commodity, "world_bank_pink_sheet", price),
                    )
                    inserted += cur.rowcount

            conn.commit()

        logger.info(
            "World Bank ingest: %d inserted, %d skipped, %d rows processed",
            inserted, skipped, rows_processed,
        )
        return {"inserted": inserted, "skipped": skipped, "rows_processed": rows_processed}
    finally:
        conn.close()
        wb.close()


def _parse_wb_date(val) -> datetime | None:
    """Parse World Bank date value (could be datetime, string like '2024M01', etc.)."""
    if isinstance(val, datetime):
        return val.replace(tzinfo=timezone.utc)

    s = str(val).strip()

    # Format: "2024M01" or "2024M1"
    if "M" in s.upper():
        parts = s.upper().split("M")
        try:
            year = int(parts[0])
            month = int(parts[1])
            return datetime(year, month, 1, tzinfo=timezone.utc)
        except (ValueError, IndexError):
            pass

    # Format: "Jan 2024" or "January 2024"
    for fmt in ("%b %Y", "%B %Y", "%Y-%m-%d", "%Y-%m", "%m/%Y"):
        try:
            return datetime.strptime(s, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue

    return None
