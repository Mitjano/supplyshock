"""CPB World Trade Monitor ingestion.

Downloads the CPB (Netherlands Bureau for Economic Policy Analysis)
World Trade Monitor Excel file and parses trade volume indices.
Stores in macro_indicators table.
Runs monthly via Celery Beat.

Issue #90
"""

import io
import logging
from datetime import datetime, timezone

import httpx
import psycopg2

from config import settings

logger = logging.getLogger(__name__)

CPB_EXCEL_URL = "https://www.cpb.nl/sites/default/files/omnidownload/CPB-World-Trade-Monitor-data.xlsx"


def ingest_cpb_trade() -> int:
    """Fetch CPB World Trade Monitor Excel and parse indices."""
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed — skipping CPB trade ingestion")
        return 0

    try:
        resp = httpx.get(CPB_EXCEL_URL, timeout=60, follow_redirects=True)
        resp.raise_for_status()
    except Exception as e:
        logger.error("CPB download failed: %s", e)
        return 0

    inserted = 0

    try:
        wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)

        # The Excel typically has a "data" or first sheet with indices
        ws = wb.active
        if ws is None:
            logger.error("CPB Excel has no active sheet")
            return 0

        # Parse headers from first row to find date columns
        headers = [cell.value for cell in ws[1]]

        # Find world trade volume column
        trade_col = None
        for i, h in enumerate(headers):
            if h and "world" in str(h).lower() and "trade" in str(h).lower():
                trade_col = i
                break

        if trade_col is None:
            # Fallback: use second column (common layout)
            trade_col = 1

        conn = psycopg2.connect(settings.DATABASE_URL_SYNC)
        try:
            with conn.cursor() as cur:
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS macro_indicators (
                        id BIGSERIAL PRIMARY KEY,
                        indicator TEXT NOT NULL,
                        time TIMESTAMPTZ NOT NULL,
                        value DOUBLE PRECISION NOT NULL,
                        unit TEXT,
                        source TEXT NOT NULL,
                        ingested_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                        UNIQUE (indicator, time, source)
                    )
                """)

                for row in ws.iter_rows(min_row=2, values_only=True):
                    date_val = row[0]
                    trade_val = row[trade_col] if trade_col < len(row) else None

                    if date_val is None or trade_val is None:
                        continue

                    # Parse date
                    if isinstance(date_val, datetime):
                        dt = date_val.replace(tzinfo=timezone.utc)
                    elif isinstance(date_val, str):
                        try:
                            dt = datetime.strptime(date_val, "%Y-%m-%d").replace(tzinfo=timezone.utc)
                        except ValueError:
                            try:
                                dt = datetime.strptime(date_val, "%Y-%m").replace(tzinfo=timezone.utc)
                            except ValueError:
                                continue
                    else:
                        continue

                    try:
                        value = float(trade_val)
                    except (ValueError, TypeError):
                        continue

                    cur.execute(
                        """
                        INSERT INTO macro_indicators
                            (indicator, time, value, unit, source)
                        VALUES (%s, %s, %s, %s, %s)
                        ON CONFLICT (indicator, time, source) DO UPDATE SET
                            value = EXCLUDED.value,
                            ingested_at = NOW()
                        """,
                        ("world_trade_volume", dt, value, "index_2010_100", "cpb"),
                    )
                    inserted += cur.rowcount

            conn.commit()
            logger.info("CPB World Trade Monitor: inserted %d records", inserted)
        finally:
            conn.close()

    except Exception as e:
        logger.error("CPB Excel parsing failed: %s", e)

    return inserted
